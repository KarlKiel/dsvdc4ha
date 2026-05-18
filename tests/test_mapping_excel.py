"""Tests for the mapping Excel generator and audit tool."""
from __future__ import annotations
import importlib.util, pathlib

_REPO = pathlib.Path(__file__).parent.parent


def _load(name, relpath):
    spec = importlib.util.spec_from_file_location(name, _REPO / relpath)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_schema_columns_have_correct_structure():
    schema = _load("excel_schema", "tools/excel_schema.py")
    assert len(schema.COLUMNS) > 30, "Expected at least 30 columns"
    headers = [h for h, _, _ in schema.COLUMNS]
    # Identity columns
    assert "domain" in headers
    assert "device_class" in headers
    assert "model" in headers
    # Enum columns
    assert "primary_group.VALUE" in headers
    assert "bi.sensor_function.USER" in headers
    assert "bi.sensor_function.VALUE" in headers
    assert "sensor.sensor_type.USER" in headers
    assert "output.function.VALUE" in headers
    assert "output.ch0.channel_type.VALUE" in headers
    assert "output.ch5.channel_type.VALUE" in headers
    assert "button.group.USER" in headers


def test_schema_enum_options_all_present():
    schema = _load("excel_schema", "tools/excel_schema.py")
    required = {
        "YesNo", "ColorGroup", "BinaryInputType", "BinaryInputGroup",
        "BinaryInputUsage", "SensorType", "SensorUsage", "SensorGroup",
        "OutputFunction", "OutputUsage", "OutputMode", "ColorClass",
        "OutputChannelType", "ButtonType", "ButtonGroup",
        "ButtonFunctionJoker", "ButtonMode",
    }
    assert required <= set(schema.ENUM_OPTIONS.keys())
    assert schema.ENUM_OPTIONS["YesNo"] == ["yes", "no"]
    assert "-" in schema.ENUM_OPTIONS["ColorGroup"]
    assert "BLACK" in schema.ENUM_OPTIONS["ColorGroup"]


def test_schema_extractors_on_known_entry():
    schema = _load("excel_schema", "tools/excel_schema.py")
    from custom_components.dsvdc4ha.entity_mapping import ENTITY_MAPPING
    col_map = {h: (h, ek, fn) for h, ek, fn in schema.COLUMNS}

    # binary_sensor/None: sensor_function_choices="any" → USER=yes, sensor_function=GENERIC
    entry = next(e for e in ENTITY_MAPPING if e["domain"] == "binary_sensor" and e["device_class"] is None)
    _, _, fn_user = col_map["bi.sensor_function.USER"]
    _, _, fn_val  = col_map["bi.sensor_function.VALUE"]
    assert fn_user(entry) == "yes"
    assert fn_val(entry) == "GENERIC"

    # binary_sensor/motion: no sensor_function_choices → USER=no, sensor_function=MOTION
    entry_m = next(e for e in ENTITY_MAPPING if e["domain"] == "binary_sensor" and e["device_class"] == "motion")
    assert fn_user(entry_m) == "no"
    assert fn_val(entry_m) == "MOTION"

    # sensor/temperature: sensor_usage_choices set → USER=yes
    entry_t = next(e for e in ENTITY_MAPPING if e["domain"] == "sensor" and e["device_class"] == "temperature")
    _, _, su_user = col_map["sensor.sensor_usage.USER"]
    assert su_user(entry_t) == "yes"

    # cover/awning: output.ch0.channel_type = SHADE_POSITION_OUTSIDE
    entry_a = next(e for e in ENTITY_MAPPING if e["domain"] == "cover" and e["device_class"] == "awning")
    _, _, ch0 = col_map["output.ch0.channel_type.VALUE"]
    assert ch0(entry_a) == "SHADE_POSITION_OUTSIDE"

    # cover/awning has no ch1 → "-"
    _, _, ch1 = col_map["output.ch1.channel_type.VALUE"]
    assert ch1(entry_a) == "-"


def test_generate_creates_file_with_correct_shape(tmp_path):
    gen = _load("generate_mapping_excel", "tools/generate_mapping_excel.py")
    schema = _load("excel_schema", "tools/excel_schema.py")
    from custom_components.dsvdc4ha.entity_mapping import ENTITY_MAPPING
    import openpyxl

    out = tmp_path / "mapping.xlsx"
    gen.generate(output_path=out)

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    assert len(rows) == len(ENTITY_MAPPING) + 1  # header + data rows
    assert rows[0][0] == "domain"
    assert rows[0][1] == "device_class"
    assert len(rows[0]) == len(schema.COLUMNS)


def test_generate_writes_correct_enum_names(tmp_path):
    gen = _load("generate_mapping_excel", "tools/generate_mapping_excel.py")
    schema = _load("excel_schema", "tools/excel_schema.py")
    import openpyxl

    out = tmp_path / "mapping.xlsx"
    gen.generate(output_path=out)

    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {v: i for i, v in enumerate(header_row) if v}

    rows = {
        (r[col_map["domain"]], r[col_map["device_class"]]): r
        for r in ws.iter_rows(min_row=2, values_only=True)
    }
    wb.close()

    # binary_sensor/None: sensor_function_choices="any" → USER=yes, VALUE=GENERIC
    row = rows[("binary_sensor", "-")]
    assert row[col_map["bi.sensor_function.USER"]] == "yes"
    assert row[col_map["bi.sensor_function.VALUE"]] == "GENERIC"
    assert row[col_map["bi.input_usage.USER"]] == "yes"
    assert row[col_map["bi.input_usage.VALUE"]] == "UNDEFINED"

    # binary_sensor/motion: no sensor_function_choices → USER=no, VALUE=MOTION
    row_m = rows[("binary_sensor", "motion")]
    assert row_m[col_map["bi.sensor_function.USER"]] == "no"
    assert row_m[col_map["bi.sensor_function.VALUE"]] == "MOTION"

    # cover/awning: ch0=SHADE_POSITION_OUTSIDE, ch1=-
    row_a = rows[("cover", "awning")]
    assert row_a[col_map["output.ch0.channel_type.VALUE"]] == "SHADE_POSITION_OUTSIDE"
    assert row_a[col_map["output.ch1.channel_type.VALUE"]] == "-"

    # sensor/temperature: sensor_usage_choices → USER=yes
    row_t = rows[("sensor", "temperature")]
    assert row_t[col_map["sensor.sensor_usage.USER"]] == "yes"

    # light/None: primary_group.USER always "no"
    row_l = rows[("light", "-")]
    assert row_l[col_map["primary_group.USER"]] == "no"


def test_generate_empty_entity_mapping(tmp_path):
    gen = _load("generate_mapping_excel", "tools/generate_mapping_excel.py")
    import openpyxl

    out = tmp_path / "empty_mapping.xlsx"
    gen.generate(entity_mapping=[], output_path=out)

    assert out.exists(), "Output file should be created even for empty entity_mapping"
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    assert len(rows) == 1, f"Expected 1 row (header only), got {len(rows)}"
    assert rows[0][0] == "domain"


def test_generate_creates_lookups_sheet(tmp_path):
    gen = _load("generate_mapping_excel", "tools/generate_mapping_excel.py")
    import openpyxl

    out = tmp_path / "mapping.xlsx"
    gen.generate(output_path=out)

    wb = openpyxl.load_workbook(out, read_only=True)
    assert "_lookups" in wb.sheetnames
    ws_lk = wb["_lookups"]
    first_col = [r[0].value for r in ws_lk.iter_rows(min_row=1, max_row=2)]
    assert "yes" in first_col
    wb.close()


def test_audit_passes_on_freshly_generated_excel(tmp_path):
    gen   = _load("generate_mapping_excel", "tools/generate_mapping_excel.py")
    audit = _load("audit_mapping",          "tools/audit_mapping.py")

    out = tmp_path / "mapping.xlsx"
    gen.generate(output_path=out)

    result = audit.run_audit(str(out))
    assert result["discrepancies"] == [], (
        "Unexpected discrepancies:\n" +
        "\n".join(str(d) for d in result["discrepancies"])
    )
    assert result["missing_entries"] == []


def test_audit_detects_wrong_value(tmp_path):
    import openpyxl
    gen   = _load("generate_mapping_excel", "tools/generate_mapping_excel.py")
    audit = _load("audit_mapping",          "tools/audit_mapping.py")
    schema = _load("excel_schema",          "tools/excel_schema.py")

    out = tmp_path / "mapping.xlsx"
    gen.generate(output_path=out)

    # Corrupt binary_sensor/motion: change bi.sensor_function.VALUE from MOTION → SMOKE
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    sf_col = schema.HEADER_INDEX["bi.sensor_function.VALUE"] + 1  # 1-based
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "binary_sensor" and row[1].value == "motion":
            row[sf_col - 1].value = "SMOKE"
            break
    wb.save(out)

    result = audit.run_audit(str(out))
    assert any(
        d["domain"] == "binary_sensor"
        and d["device_class"] == "motion"
        and d["field"] == "sensor_function"
        for d in result["discrepancies"]
    ), f"Expected discrepancy not in: {result['discrepancies']}"


def test_audit_detects_wrong_user_flag(tmp_path):
    import openpyxl
    gen   = _load("generate_mapping_excel", "tools/generate_mapping_excel.py")
    audit = _load("audit_mapping",          "tools/audit_mapping.py")
    schema = _load("excel_schema",          "tools/excel_schema.py")

    out = tmp_path / "mapping.xlsx"
    gen.generate(output_path=out)

    # Corrupt binary_sensor/door: change bi.group.USER from "no" to "yes"
    # (door has no group_choices, so "yes" is wrong)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    gu_col = schema.HEADER_INDEX["bi.group.USER"] + 1
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "binary_sensor" and row[1].value == "door":
            row[gu_col - 1].value = "yes"
            break
    wb.save(out)

    result = audit.run_audit(str(out))
    assert any(
        d["domain"] == "binary_sensor"
        and d["device_class"] == "door"
        and "group" in d["field"]
        and "USER" in d["field"]
        for d in result["discrepancies"]
    ), f"Expected USER discrepancy not in: {result['discrepancies']}"
