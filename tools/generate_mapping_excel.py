"""Generate ha_vdsd_mapping.xlsx from entity_mapping.py and pydsvdcapi enums.

Usage:
    python tools/generate_mapping_excel.py
    python tools/generate_mapping_excel.py path/to/output.xlsx
"""
from __future__ import annotations
import importlib.util, pathlib, sys

_REPO_ROOT = pathlib.Path(__file__).parent.parent
sys.path.insert(0, str(_REPO_ROOT))

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from tools.excel_schema import COLUMNS, ENUM_OPTIONS


def _load_entity_mapping():
    spec = importlib.util.spec_from_file_location(
        "entity_mapping",
        _REPO_ROOT / "custom_components/dsvdc4ha/entity_mapping.py",
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.ENTITY_MAPPING


_FILL_HEADER = PatternFill("solid", fgColor="4472C4")   # blue
_FILL_USER   = PatternFill("solid", fgColor="E2EFDA")   # light green
_FONT_HEADER = Font(color="FFFFFF", bold=True)
_FONT_USER_H = Font(color="375623", bold=True)


def _write_lookups(wb: openpyxl.Workbook) -> dict[str, str]:
    """Create hidden _lookups sheet; return {enum_key: range_ref} for DataValidation."""
    ws = wb.create_sheet("_lookups")
    ws.sheet_state = "hidden"
    refs: dict[str, str] = {}
    for col_idx, (key, options) in enumerate(ENUM_OPTIONS.items(), 1):
        col_l = get_column_letter(col_idx)
        for row_idx, opt in enumerate(options, 1):
            ws.cell(row=row_idx, column=col_idx, value=opt)
        refs[key] = f"_lookups!${col_l}$1:${col_l}${len(options)}"
    return refs


def generate(
    entity_mapping=None,
    output_path: str | pathlib.Path = "documents/ha_vdsd_mapping.xlsx",
) -> pathlib.Path:
    """Generate the mapping Excel file. Pass entity_mapping to override (useful in tests)."""
    if entity_mapping is None:
        entity_mapping = _load_entity_mapping()
    output_path = pathlib.Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapping"

    lookup_refs = _write_lookups(wb)

    # Header row
    for col_idx, (header, enum_key, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if enum_key == "YesNo":
            cell.fill = _FILL_USER
            cell.font = _FONT_USER_H
        else:
            cell.fill = _FILL_HEADER
            cell.font = _FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    # Data rows
    n_rows = len(entity_mapping)
    for row_idx, entry in enumerate(entity_mapping, 2):
        for col_idx, (_, _, extractor) in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=extractor(entry))

    # DataValidation dropdowns — one DV per column that has an enum_key
    for col_idx, (_, enum_key, _) in enumerate(COLUMNS, 1):
        if enum_key is None:
            continue
        col_l = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=lookup_refs[enum_key],
            allow_blank=True,
            showErrorMessage=False,
        )
        dv.sqref = f"{col_l}2:{col_l}{n_rows + 1}"
        ws.add_data_validation(dv)

    # Column widths
    for col_idx, (header, _, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 14)

    wb.save(output_path)
    print(f"Generated {output_path} ({n_rows} rows, {len(COLUMNS)} columns)")
    return output_path


def main(argv=None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    path = args[0] if args else "documents/ha_vdsd_mapping.xlsx"
    generate(output_path=path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
